from natsort import natsorted
from openpyxl import Workbook
from openpyxl.styles import Alignment
import json
import os
import sys

import rules

#FUNCTIONS

#returns a list of controls given a JSON file containing a security template
def getControls(templatePath, elementRules, areaRules):
    #initialize list for controls
    controlList = []
    #load template
    with open(templatePath) as templateFile:
        templateData = json.load(templateFile)
    #apply rules for elements
    for rule in elementRules:
        controlList.extend(rule(templateData))
    #check if there are areas
    if len(templateData) == 2:
        #there are no areas, close the file and return
        templateFile.close()
        #print log information
        print('Processed: ' + templatePath + ' , number of controls: '+str(len(controlList)))
        #if no control were extracted put placeholder
        if len(controlList)<1:
            controlList.append("N/A")
        #return list of controls
        return controlList
    #there are areas apply rules
    for rule in areaRules:
        controlList.extend(rule(templateData))
    #print log information
    print('Processed: ' + templatePath + ' , number of controls: '+str(len(controlList)))
    #if no control were extracted put placeholder
    if len(controlList)<1:
        controlList.append("N/A")
    #return list of controls
    return controlList

#prints all controls of a subcategory in a spreadsheets, returns the next free row
def exportSubcategory(templatePath, worksheet, firstRow, elementRules, areaRules, skip):
    #get all controls from subcategory
    controls = getControls(templatePath, elementRules, areaRules)
    #initialize index for controls
    i = 0
    #append them in the correct column
    for control in controls:
        if not skip:
            control = 'CTR-'+templatePath.split('/')[-1].strip(".JSON")+'-'+str(i)+": "+control
        worksheet.append(['','','',control])
        i += 1
    #put the name of the subcategory in the corret colum in the first row
    worksheet['C'+str(firstRow)] = templatePath.split('/')[-1].strip(".JSON")
    #allign text
    c = worksheet['C'+str(firstRow)]
    c.alignment = Alignment(vertical='center')
    #compute the last row
    nextRow = firstRow + len(controls)
    #merge the cells of the subcategory name
    worksheet.merge_cells('C'+str(firstRow)+':C'+str(nextRow-1))
    #return the first new free row in the worksheet
    return nextRow


#prints all subcategories of a catagory in a spreadsheet, returns the next free row
def exportCategory(categoryDirectory, worksheet, firstRow, elementRules, areaRules, skip):
    #gets all subcategory from category
    templates = os.listdir(categoryDirectory)
    templates = natsorted(templates)
    #initialize row pointer
    nextRow = firstRow
    #for each subcategory
    for template in templates:
        #build the path
        templatePath = categoryDirectory + '/' + template
        #export the subcategory
        nextRow = exportSubcategory(templatePath, worksheet, nextRow, elementRules, areaRules, skip)
    #put the name of the category in the correct cell
    worksheet['B'+str(firstRow)] = categoryDirectory.split('/')[-1]
    #allign text
    c = worksheet['B'+str(firstRow)]
    c.alignment = Alignment(vertical='center')
    #merge the cells of the category name
    worksheet.merge_cells('B'+str(firstRow)+':B'+str(nextRow-1))
    #return the next free row
    return nextRow

#prints all categories of a function in a spreadsheet, returns the next free row
def exportFunction(functionDirectory, worksheet, firstRow, elementRules, areaRules, skip):
    #get all categories from the function
    categories = os.listdir(functionDirectory)
    categories = natsorted(categories)
    #initalize the row pointer
    nextRow = firstRow
    #for each category
    for category in categories:
        #build the path
        categoryPath = functionDirectory + '/' + category
        #export the category
        nextRow = exportCategory(categoryPath, worksheet, nextRow, elementRules, areaRules, skip)
    #put the name of the function in the correct cell
    worksheet['A'+str(firstRow)] = functionDirectory.split('/')[-1]
    #allign text
    c = worksheet['A'+str(firstRow)]
    c.alignment = Alignment(vertical='center')
    #merge the cells of the function name
    worksheet.merge_cells('A'+str(firstRow)+':A'+str(nextRow-1))
    #return the next free row
    return nextRow



#MAIN

#retrive name for outputfile
try:
    templatesPath = sys.argv[1]
    outputFile = sys.argv[2]
except IndexError:
    raise SystemExit(f"Usage: {sys.argv[0]} <security_templates_path> <output_xlsx_path> OPTIONAL: -n")

try:
    if 'n' in sys.argv[3] : skipIds = True
except IndexError:
    skipIds = False



#initialize array of rules
elementRules = []
areaRules = []
#load config file
with open("config.JSON") as configFile:
    configData = json.load(configFile)
#append elements rules
if configData["nodeExist"] == True: elementRules.append(rules.nodeExist)
if configData["nodeC"] == True: elementRules.append(rules.nodeC)
if configData["nodeI"] == True: elementRules.append(rules.nodeI)
if configData["nodeA"] == True: elementRules.append(rules.nodeA)
if configData["nodeCIA"] == True: elementRules.append(rules.nodeCIA)
if configData["flowExist"] == True: elementRules.append(rules.flowExist)
if configData["flowC"] == True: elementRules.append(rules.flowC)
if configData["flowI"] == True: elementRules.append(rules.flowI)
if configData["flowA"] == True: elementRules.append(rules.flowA)
if configData["flowP"] == True: elementRules.append(rules.flowP)
if configData["flowCIAP"] == True: elementRules.append(rules.flowCIAP)
#append area rules
if configData["areaExist"] == True: areaRules.append(rules.areaExist)
if configData["areaTrust"] == True: areaRules.append(rules.areaTrust)
if configData["areaFlowsEnter"] == True: areaRules.append(rules.areaFlowsEnter)
if configData["areaFlowsExit"] == True: areaRules.append(rules.areaFlowsExit)


#initialize list of functions names so they are in the correct order
functions = ['Identify', 'Protect', 'Detect', 'Respond', 'Recover']
#create a new workbook and worksheet on it
wb = Workbook()
ws = wb.active
#initialize colums names
ws.append(['Function', 'Category', 'Subcategory', 'Controls'])
#initialize nextRow
nextRow = 2
#for each function
for function in functions:
    #build the path
    functionPath = templatesPath+function
    #export function
    if os.path.isdir(functionPath):
        nextRow = exportFunction(functionPath, ws, nextRow, elementRules, areaRules, skipIds)
#save before exiting
wb.save(outputFile)


#TEST 1 : PASS
#controls = getControls('Security Templates/Identify Function/ID.AM/ID.AM-6.JSON')
#for c in controls:
#   print(c)

#TEST 2 : PASS
#create a new workbook and worksheet on it
#wb = Workbook()
#ws = wb.active
#initialize colums names
#ws.append(['Function', 'Category', 'Subcategory', 'Controls'])
#print subcategory in the worksheet
#nextRow = exportSubcategory('Security Templates/Identify/ID.AM/ID.AM-6.JSON', ws, 2)
#print the next row and save
#print(nextRow)
#wb.save("test.xlsx")

#TEST 3 : PASS
#create a new workbook and worksheet on it
#wb = Workbook()
#ws = wb.active
#initialize colums names
#ws.append(['Function', 'Category', 'Subcategory', 'Controls'])
#print a category in the worksheet
#nextRow = exportCategory('Security Templates/Protect/PR.IP', ws, 2)
#print the next row and save
#print(nextRow)
#wb.save("test.xlsx")

#TEST 4 : PASS
#create a new workbook and worksheet on it
#wb = Workbook()
#ws = wb.active
#initialize colums names
#ws.append(['Function', 'Category', 'Subcategory', 'Controls'])
#print a function in the worksheet
#nextRow = exportFunction('Security Templates/Protect', ws, 2)
#print the next row and save
#print(nextRow)
#wb.save("test.xlsx")